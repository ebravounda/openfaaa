import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EUR_FMT = '#,##0.00 €'
HEADER_FILL = PatternFill("solid", fgColor="0052FF")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT = Font(bold=True, size=14, color="0F172A")
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _repercutido_rows(invoices):
    rows = []
    for inv in invoices:
        rows.append([
            inv.get("issue_date", ""), inv.get("number", ""),
            inv.get("client", {}).get("name", ""), inv.get("client", {}).get("nif", ""),
            inv.get("base", 0), inv.get("iva_rate", 0), inv.get("iva_amount", 0),
            inv.get("irpf_rate", 0), inv.get("irpf_amount", 0), inv.get("total", 0),
        ])
    return rows


def _soportado_rows(expenses):
    rows = []
    for exp in expenses:
        rows.append([
            exp.get("date", ""), exp.get("vendor_name", ""), exp.get("vendor_nif", ""),
            exp.get("category", ""), exp.get("base", 0), exp.get("iva_rate", 0),
            exp.get("iva_amount", 0), exp.get("total", 0),
        ])
    return rows


def build_libros_xlsx(company, invoices, expenses, year) -> bytes:
    wb = Workbook()

    # --- IVA Repercutido ---
    ws1 = wb.active
    ws1.title = "IVA Repercutido"
    ws1["A1"] = f"Libro de IVA Repercutido {year} — {company.get('name', '')}"
    ws1["A1"].font = TITLE_FONT
    headers1 = ["Fecha", "Nº Factura", "Cliente", "NIF/CIF", "Base", "% IVA",
                "Cuota IVA", "% IRPF", "Ret. IRPF", "Total"]
    ws1.append([])
    ws1.append(headers1)
    _style_header(ws1, 3, len(headers1))
    for r in _repercutido_rows(invoices):
        ws1.append(r)
    last = ws1.max_row
    for row in range(4, last + 1):
        for col in (5, 7, 9, 10):
            ws1.cell(row=row, column=col).number_format = EUR_FMT
    ws1.append(["", "", "", "TOTALES",
                sum(i.get("base", 0) for i in invoices), "",
                sum(i.get("iva_amount", 0) for i in invoices), "",
                sum(i.get("irpf_amount", 0) for i in invoices),
                sum(i.get("total", 0) for i in invoices)])
    tr = ws1.max_row
    for col in (4, 5, 7, 9, 10):
        ws1.cell(row=tr, column=col).font = TOTAL_FONT
        if col in (5, 7, 9, 10):
            ws1.cell(row=tr, column=col).number_format = EUR_FMT
    _autosize(ws1, [12, 14, 28, 14, 14, 8, 14, 8, 14, 14])

    # --- IVA Soportado ---
    ws2 = wb.create_sheet("IVA Soportado")
    ws2["A1"] = f"Libro de IVA Soportado {year} — {company.get('name', '')}"
    ws2["A1"].font = TITLE_FONT
    headers2 = ["Fecha", "Proveedor", "NIF/CIF", "Categoría", "Base", "% IVA", "Cuota IVA", "Total"]
    ws2.append([])
    ws2.append(headers2)
    _style_header(ws2, 3, len(headers2))
    for r in _soportado_rows(expenses):
        ws2.append(r)
    last2 = ws2.max_row
    for row in range(4, last2 + 1):
        for col in (5, 7, 8):
            ws2.cell(row=row, column=col).number_format = EUR_FMT
    ws2.append(["", "", "", "TOTALES",
                sum(e.get("base", 0) for e in expenses), "",
                sum(e.get("iva_amount", 0) for e in expenses),
                sum(e.get("total", 0) for e in expenses)])
    tr2 = ws2.max_row
    for col in (4, 5, 7, 8):
        ws2.cell(row=tr2, column=col).font = TOTAL_FONT
        if col in (5, 7, 8):
            ws2.cell(row=tr2, column=col).number_format = EUR_FMT
    _autosize(ws2, [12, 28, 14, 14, 14, 8, 14, 14])

    # --- Resumen ---
    ws3 = wb.create_sheet("Resumen")
    ws3["A1"] = f"Resumen fiscal {year}"
    ws3["A1"].font = TITLE_FONT
    iva_rep = sum(i.get("iva_amount", 0) for i in invoices)
    iva_sop = sum(e.get("iva_amount", 0) for e in expenses)
    irpf = sum(i.get("irpf_amount", 0) for i in invoices)
    ingresos = sum(i.get("base", 0) for i in invoices)
    gastos = sum(e.get("base", 0) for e in expenses)
    summary = [
        ("Ingresos (base imponible)", ingresos),
        ("Gastos (base imponible)", gastos),
        ("Beneficio", ingresos - gastos),
        ("IVA Repercutido", iva_rep),
        ("IVA Soportado", iva_sop),
        ("IVA a pagar (303)", iva_rep - iva_sop),
        ("IRPF retenido en facturas", irpf),
    ]
    ws3.append([])
    for label, val in summary:
        ws3.append([label, val])
        ws3.cell(row=ws3.max_row, column=2).number_format = EUR_FMT
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
    _autosize(ws3, [30, 18])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_libros_csv(invoices, expenses, year) -> str:
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow(["Libro", "Fecha", "Documento", "Contraparte", "NIF/CIF",
                "Base", "% IVA", "Cuota IVA", "% IRPF", "Ret. IRPF", "Total"])
    for inv in invoices:
        w.writerow(["Repercutido", inv.get("issue_date", ""), inv.get("number", ""),
                    inv.get("client", {}).get("name", ""), inv.get("client", {}).get("nif", ""),
                    inv.get("base", 0), inv.get("iva_rate", 0), inv.get("iva_amount", 0),
                    inv.get("irpf_rate", 0), inv.get("irpf_amount", 0), inv.get("total", 0)])
    for exp in expenses:
        w.writerow(["Soportado", exp.get("date", ""), exp.get("category", ""),
                    exp.get("vendor_name", ""), exp.get("vendor_nif", ""),
                    exp.get("base", 0), exp.get("iva_rate", 0), exp.get("iva_amount", 0),
                    0, 0, exp.get("total", 0)])
    return "\ufeff" + out.getvalue()
